import openpyxl
from telebot import types
from keyboards import massage_types_markup, time_slots_markup
from menu import main_menu
import datetime
from config import bot, doktor_id
def register_massage(message):
    # Запрашиваем данные от пользователя
    bot.send_message(message.chat.id, "Введите ФИО:")
    bot.register_next_step_handler(message, ask_phone_number)

def ask_phone_number(message):
    # Сохраняем ФИО и запрашиваем номер телефона
    fio = message.text
    bot.send_message(message.chat.id, "Введите номер телефона:")
    bot.register_next_step_handler(message, ask_massage_type, fio)

def ask_massage_type(message, fio):
    # Сохраняем номер телефона и предлагаем выбрать тип массажа
    phone_number = message.text
    bot.send_message(message.chat.id, "Выберите тип массажа:", reply_markup=massage_types_markup())
    bot.register_next_step_handler(message, ask_date, fio, phone_number)

def ask_date(message, fio, phone_number):
    # Сохраняем тип массажа и предлагаем ввести дату
    massage_type = message.text
    bot.send_message(message.chat.id, "Введите дату (в формате дд.мм.гггг):")
    bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)

def ask_time(message, fio, phone_number, massage_type):
    date = message.text

    if not is_valid_date(date):
        bot.send_message(message.chat.id, "Вы ввели неверную дату. Введите дату начиная с сегодняшней.")
        bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)
        return

    entered_date = datetime.datetime.strptime(date, "%d.%m.%Y").date()
    today = datetime.date.today()

    if entered_date < today:
        bot.send_message(message.chat.id, "На прошедшую дату запись невозможна.")
        bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)
        return

    if is_weekend(entered_date):
        bot.send_message(message.chat.id, "Введённая дата является выходным днем. Запишитесь на будний день.")
        bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)
        return

    if check_date_format(date):
        available_times = get_available_times(date)

        if available_times:
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            for time in available_times:
                keyboard.add(time)

            bot.send_message(message.chat.id, "Выберите удобное время:", reply_markup=keyboard)
            bot.register_next_step_handler(message, confirm_registration, fio, phone_number, massage_type, date)
        else:
            bot.send_message(message.chat.id, "К сожалению, нет свободного времени на выбранную дату. Выберите другой день.")
            bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)
    else:
        bot.send_message(message.chat.id, "Некорректный формат даты. Введите дату (в формате дд.мм.гггг):")
        bot.register_next_step_handler(message, ask_time, fio, phone_number, massage_type)

def confirm_registration(message, fio, phone_number, massage_type, date):
    # Подтверждаем запись и сохраняем данные в файл db.xlsx
    time_slot = message.text
    bot.send_message(message.chat.id, "Подтвердите запись:\n\n"
                                      f"ФИО: {fio}\n"
                                      f"Номер телефона: {phone_number}\n"
                                      f"Тип массажа: {massage_type}\n"
                                      f"Дата: {date}\n"
                                      f"Время: {time_slot}",
                     reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True).add(types.KeyboardButton("Да"),
                                                                                        types.KeyboardButton("Нет")))
    bot.register_next_step_handler(message, save_registration, fio, phone_number, massage_type, date, time_slot)

def save_registration(message, fio, phone_number, massage_type, date, time_slot):
    # Сохраняем данные о записи в файл db.xlsx
    if message.text.lower() == 'да':
        wb = openpyxl.load_workbook('db.xlsx')
        sheet = wb.active
        row = [fio, phone_number, massage_type, date, time_slot]
        sheet.append(row)
        wb.save('db.xlsx')
        bot.send_message(doktor_id, "записан на массаж:\n\n"
                                      f"ФИО: {fio}\n"
                                      f"Номер телефона: {phone_number}\n"
                                      f"массаж: {massage_type}\n"
                                      f"Дата: {date}\n"
                                      f"Время: {time_slot}")
        bot.send_message(message.chat.id, "Ваша запись сохранена. Ждем вас на массаж!")
        main_menu(message)

    else:
        register_massage(message)

def get_available_times(date):
    wb = openpyxl.load_workbook('db.xlsx')
    sheet = wb.active
    times = set()

    for row in sheet.iter_rows(values_only=True):
        if row[3] == date:
            times.add(row[4])

    all_times = {'09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30'}
    available_times = all_times - times

    return available_times


def is_valid_date(date):
    try:
        datetime.datetime.strptime(date, "%d.%m.%Y")
        return True
    except ValueError:
        return False

def is_weekend(date):
    return date.weekday() >= 5

def check_date_format(date):
    try:
        datetime.datetime.strptime(date, '%d.%m.%Y')
        return True
    except ValueError:
        return False
